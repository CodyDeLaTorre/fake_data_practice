from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
fake_data = Faker()

# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())

for i in range(1,11):
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()
    ws.cell(row=i, column=3).value = fake_data.address()

wb.save("test_data.xlsx")

